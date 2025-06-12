import os

def build_file_path(*path_components):
    """
    构建文件路径，支持传入多个路径组件
    
    参数:
    *path_components (str): 可变长度的路径组件参数
    
    返回:
    str: 连接后的文件路径
    """
    return os.path.join(*path_components)

from openpyxl import Workbook

def create_excel_file(file_name):
    """
    创建一个新的 Excel 文件。

    参数:
    file_name (str): 要创建的 Excel 文件的名称，需包含 .xlsx 扩展名。

    返回:
    bool: 如果文件创建成功返回 True，失败返回 False。
    """
    try:
        # 创建一个新的工作簿
        wb = Workbook()
        # 保存工作簿到指定文件
        wb.save(file_name)
        return True
    except Exception as e:
        print(f"创建 Excel 文件时出错: {e}")
        return False
    
from openpyxl import Workbook

def create_new_worksheet(workbook, sheet_name):
    """
    在指定的工作簿中创建一个新的工作表。

    参数:
    workbook (openpyxl.workbook.workbook.Workbook): 要操作的工作簿对象。
    sheet_name (str): 新工作表的名称。

    返回:
    openpyxl.worksheet.worksheet.Worksheet: 新创建的工作表对象，如果创建失败则返回 None。
    """
    try:
        # 在工作簿中创建一个新的工作表
        new_sheet = workbook.create_sheet(title=sheet_name)
        return new_sheet
    except Exception as e:
        print(f"创建工作表 {sheet_name} 时出错: {e}")
        return None
    
import os
from PIL import Image

def check_image_integrity(folder_path):
    """
    遍历指定文件夹中的所有图片，检查图片是否损坏。

    参数:
    folder_path (str): 要检查的文件夹路径。

    返回:
    tuple: 包含三个元素，
           1. 该文件夹中图片的数量；
           2. 损坏图片的数量；
           3. 损坏图片的文件路径列表。
    """
    image_extensions = ('.jpg', '.jpeg', '.png', '.gif', '.bmp')
    total_image_count = 0
    corrupted_image_count = 0
    corrupted_image_paths = []

    if os.path.exists(folder_path):
        for root, dirs, files in os.walk(folder_path):
            for file in files:
                if file.lower().endswith(image_extensions):
                    total_image_count += 1
                    file_path = os.path.join(root, file)
                    try:
                        with Image.open(file_path) as img:
                            img.verify()
                    except Exception:
                        corrupted_image_count += 1
                        corrupted_image_paths.append(file_path)

    return total_image_count, corrupted_image_count, corrupted_image_paths


if __name__ == "__main__":
    # 示例用法
    print("-" * 50)  # 分隔线
    '''
    # 构建文件路径
    file_path = build_file_path("folder", "subfolder", "file.txt")
    # 构建文件路径
    file_path = build_file_path(folder_path, 'file.txt')
    print(f"构建的文件路径: {file_path}")
    # 构建一个指向文件的路径
    config_path = build_file_path("config", "settings.ini")
    # 输出: config/settings.ini（Linux/macOS）或 config/settings.ini（Windows）
    # 构建多级目录路径
    data_path = build_file_path("home", "user", "data", "file.txt")
    # 输出: home/user/data/file.txt（Linux/macOS）或 home/user/data/file.txt（Windows）
    # 构建相对路径
    relative_path = build_file_path("..", "documents", "file.txt")
    # 输出: ..\\documents\\file.txt（Linux/macOS）或 ..\\documents\file.txt（Windows）
'''
    print("-" * 50)  # 分隔线
    '''
    # 指定要创建的 Excel 文件的名称
    file_name = "example.xlsx"
    # 创建一个新的 Excel 文件
    excel_file_name = 'example.xlsx'
    if create_excel_file(excel_file_name):
        print(f"Excel 文件 '{excel_file_name}' 创建成功。")
    else:
        print(f"Excel 文件 '{excel_file_name}' 创建失败。")
    if create_excel_file(file_name):
        print(f"{file_name} 文件创建成功。")
        wb = Workbook()
        new_sheet = create_new_worksheet(wb, "NewSheet")
        if new_sheet:
            print(f"工作表 {new_sheet.title} 创建成功。")
            wb.save(file_name)
    else:
        print(f"{file_name} 文件创建失败。")
'''
    print("-" * 50)  # 分隔线
    '''
    folder_path = "your_folder_path"
    total_images, corrupted_images, corrupted_paths = check_image_integrity(folder_path)
    print(f"该文件夹中图片的数量: {total_images}")
    print(f"损坏图片的数量: {corrupted_images}")
    print("损坏图片的文件路径:")
    for path in corrupted_paths:
        print(path)
'''
    print("-" * 50)  # 分隔线
